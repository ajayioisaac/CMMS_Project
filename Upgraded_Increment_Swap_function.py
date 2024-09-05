# -*- coding: utf-8 -*-
"""
Created on Sun Dec 10 09:11:54 2023

@author: Planner
"""
from openpyxl import load_workbook
import pandas as pd
from dateutil.relativedelta import relativedelta 

def add_week(input_date, n):
    # Check for NaN values before performing the calculation
    if pd.notna(input_date):
        return input_date + relativedelta(weeks=int(n))
    else:
        return None

def add_month(input_date, n):
    # Check for NaN values before performing the calculation
    if pd.notna(input_date):
        return input_date + relativedelta(months=int(n))
    else:
        return None

def swap(sheet):
    # Iterate through rows starting from the second row (assuming headers are in the first row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        last_due = row[9].value  # Assuming the last_due column is in the 10th column (index 9)
        last_done = row[8].value  # Assuming the interval column is in the 7th column (index 6)
        
        # Convert last_due and last_done to datetime
        last_due = pd.to_datetime(last_due, errors='coerce')
        last_done = pd.to_datetime(last_done, errors='coerce')

        if not pd.isna(last_done) and pd.isna(last_due):
            last_due = last_done
            # Update the next_due column with the calculated value
            sheet.cell(row=row[9].row, column=10, value=last_due)  # Assuming the last_due column is in the 10th column (index 10)
            
        elif not pd.isna(last_done) and not pd.isna(last_due) and last_done  > last_due:
            last_due = last_done
            # Update the next_due column with the calculated value
            sheet.cell(row=row[9].row, column=10, value=last_due)  # Assuming the last_due column is in the 10th column (index 10)

def increment(sheet):
    # Iterate through rows starting from the second row (assuming headers are in the first row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        occurrence = row[7].value
        last_due = row[9].value  # Assuming the last_due column is in the 10th column (index 9)
        interval = row[6].value  # Assuming the interval column is in the 7th column (index 6)

        if occurrence == "Week":
            # Perform the incremental date function
            next_due = add_week(last_due, interval)
        elif occurrence == "Month":
            # Perform the incremental date function
            next_due = add_month(last_due, interval)
        # Update the next_due column with the calculated value
        sheet.cell(row=row[10].row, column=11, value=next_due)  # Assuming the next_due column is in the 11th column (index 10)

def clean_swap():
    # Load the workbook
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # Perform the swap function
    swap(sheet)

    # Save the changes to the Excel file
    workbook.save(excel_file_path)

    # Perform the increment function
    increment(sheet)

    # Save the changes to the Excel file
    workbook.save(excel_file_path)

    # Close the workbook
    workbook.close()

# Specify the file path of the Excel sheet
excel_file_path = 'C:/Users/Planner/Documents/Planning/Working Document.xlsx'
